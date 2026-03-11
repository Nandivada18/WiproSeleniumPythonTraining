import openpyxl


def get_excel_data(file_path, sheet_name):
    # open the work book
    workbook = openpyxl.load_workbook(file_path)
    # fetch  the sheet name
    sheet = workbook[sheet_name]

    data = []

    for row in range(2, sheet.max_row+1 ):
        Username = sheet.cell(row=row, column=1).value
        Password = sheet.cell(row=row, column=2).value
        Product_Name= sheet.cell(row=row, column=3).value
        First_Name = sheet.cell(row=row, column=4).value
        Last_Name = sheet.cell(row=row, column=5).value
        Address = sheet.cell(row=row, column=6).value
        State = sheet.cell(row=row, column=7).value
        Postal_code = sheet.cell(row=row, column=8).value
        data.append((Username, Password,Product_Name,First_Name,Last_Name,Address,State,Postal_code))

    return data
